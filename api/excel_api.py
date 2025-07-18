#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel相關API模組
包含Excel檔案處理、格式驗證、匯出等功能
"""

import os
import tempfile
import shutil
from typing import Dict, List, Any, Optional
from datetime import datetime
from io import BytesIO

import polars as pl
from fastapi import HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.utils import Utils

def print_log(message: str):
    """日誌輸出"""
    print(f"********** excel_api.py - {message}")

class ExcelAPI:
    """Excel API類別"""
    
    @staticmethod
    async def validate_excel_format(excel_file: UploadFile, required_columns: List[str]) -> pl.DataFrame:
        """
        驗證Excel檔案格式
        
        Args:
            excel_file: 上傳的Excel檔案
            required_columns: 必需的欄位列表
            
        Returns:
            標準化後的DataFrame
        """
        try:
            # 處理檔案上傳
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            shutil.copyfileobj(excel_file.file, temp_file)
            temp_file.flush()
            temp_file.close()

            try:
                # 讀取Excel檔案
                excel_data = Utils.read_excel_file(temp_file.name)
                if excel_data.is_empty():
                    raise HTTPException(status_code=400, detail="Excel檔案為空或格式錯誤")
                # 使用 Utils.standardize_columns 進行欄位標準化
                try:
                    excel_data = Utils.standardize_columns(excel_data, required_columns)
                except ValueError as e:
                    # 如果標準化失敗，嘗試只標準化股票代碼欄位
                    try:
                        excel_data = Utils.standardize_columns(excel_data, ["stock_id"])
                    except ValueError:
                        raise HTTPException(status_code=400, detail=f"Excel檔案欄位標準化失敗: {str(e)}")
                
                return excel_data
                
            finally:
                # 清理臨時檔案
                os.unlink(temp_file.name)
                
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel檔案處理失敗: {str(e)}")
    
    @staticmethod
    async def process_stock_source_excel(excel_file: UploadFile) -> Dict[str, Any]:
        """
        處理股票來源Excel檔案
        
        Args:
            excel_file: 上傳的Excel檔案
            
        Returns:
            包含股票代碼列表和Excel資料的字典
        """
        try:
            # 驗證Excel格式
            excel_data = await ExcelAPI.validate_excel_format(excel_file, ["stock_id", "date"])   
            
            # 取得股票代碼列表
            if "stock_id" not in excel_data.columns:
                raise HTTPException(status_code=400, detail="Excel檔案中找不到股票代碼欄位")
            
            stock_ids = excel_data["stock_id"].unique().to_list()
            stock_ids = [s.split(' ')[0].strip() for s in stock_ids]
            
            return {
                "excel_data": excel_data,
                "stock_ids": stock_ids
            }
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"股票來源Excel處理失敗: {str(e)}")
    
    @staticmethod
    async def process_price_excel(excel_file: UploadFile, stock_ids: Optional[List[str]] = None, 
                                start_date: Optional[str] = None, end_date: Optional[str] = None) -> pl.DataFrame:
        """
        處理股價資料Excel檔案
        
        Args:
            excel_file: 上傳的Excel檔案
            stock_ids: 股票代碼列表（可選）
            start_date: 開始日期（可選）
            end_date: 結束日期（可選）
            
        Returns:
            標準化後的股價資料DataFrame
        """
        try:
            # 驗證Excel格式
            required_columns = ["stock_id", "date", "open", "high", "low", "close"]
            price_data = await ExcelAPI.validate_excel_format(excel_file, required_columns)
            # 將股票代碼轉換為字串
            price_data = price_data.with_columns([
                pl.col("stock_id").cast(pl.Utf8)
            ])
            # 過濾指定股票和日期範圍的資料
            if stock_ids is not None:
                price_data = price_data.filter(pl.col("stock_id").is_in(stock_ids))
            if start_date and end_date:
                price_data = price_data.filter(
                    (pl.col("date") >= start_date) & (pl.col("date") <= end_date)
                )
            
            return price_data
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"股價資料Excel處理失敗: {str(e)}")
    
    @staticmethod
    async def export_backtest_excel(trade_records: List[Dict], filename: str = None) -> FileResponse:
        """
        匯出回測結果為Excel
        
        Args:
            trade_records: 交易記錄列表
            filename: 檔案名稱（可選）
            
        Returns:
            Excel檔案回應
        """
        try:
            if not trade_records:
                raise HTTPException(status_code=400, detail="沒有交易記錄可匯出")
            
            # 轉換為DataFrame
            export_data = []
            for trade in trade_records:
                export_data.append({
                    "進場日期": trade.get("entry_date", ""),
                    "出場日期": trade.get("exit_date", ""),
                    "股票代碼": trade.get("stock_id", ""),
                    "股票名稱": trade.get("stock_name", ""),
                    "交易方向": trade.get("trade_direction", ""),
                    "進場價": trade.get("entry_price", 0),
                    "出場價": trade.get("exit_price", 0),
                    "股數": trade.get("shares", 0),
                    "損益": trade.get("profit_loss", trade.get("profit", 0)),
                    "損益率": trade.get("profit_loss_rate", trade.get("profit_rate", 0)),
                    "手續費": trade.get("commission", 0),
                    "證交稅": trade.get("securities_tax", 0),
                    "淨損益": trade.get("net_profit_loss", trade.get("profit", 0)),
                    "出場原因": trade.get("exit_reason", trade.get("exit_status", "")),
                    "持有天數": trade.get("holding_days", 0)
                })
            
            df = pl.DataFrame(export_data)
            
            # 產生檔案名
            if not filename:
                filename = Utils.generate_unique_filename("backtest_result", "xlsx")
            filepath = os.path.join("data/exports", filename).replace('\\', '/')
            
            # 儲存檔案
            Utils.save_excel_file(df, filepath, "回測結果")
            
            return FileResponse(
                filepath,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=filename
            )
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"匯出Excel失敗: {str(e)}")
    
    @staticmethod
    async def export_detailed_records(
        strategy_name: str,
        complete_trade_records: List[Dict],
        complete_holding_positions: List[Dict]
    ) -> StreamingResponse:
        """
        匯出詳細回測記錄為Excel
        
        Args:
            strategy_name: 策略名稱
            complete_trade_records: 完整交易記錄
            complete_holding_positions: 完整持有部位
            
        Returns:
            Excel檔案回應
        """
        try:
            if not complete_trade_records and not complete_holding_positions:
                raise HTTPException(status_code=400, detail="沒有記錄可匯出")
            
            # 準備交易記錄資料
            trade_data = []
            for trade in complete_trade_records:
                trade_data.append({
                    "進場日期": trade.get("entry_date", ""),
                    "出場日期": trade.get("exit_date", ""),
                    "股票代碼": trade.get("stock_id", ""),
                    "股票名稱": trade.get("stock_name", ""),
                    "交易方向": "買入" if trade.get("trade_direction", 1) == 1 else "賣出",
                    "進場價格": trade.get("entry_price", 0),
                    "當日出場價": trade.get("exit_price", 0),
                    "當日進場價": trade.get("current_price", 0),                    
                    "股數": trade.get("shares", 0),
                    "報酬": trade.get("profit_loss", 0),
                    "報酬率(%)": trade.get("profit_loss_rate", 0),
                    "淨損益": trade.get("net_profit_loss", 0),                     
                    "出場價類型": trade.get("exit_price_type", ""),
                    "手續費": trade.get("commission", 0),
                    "證交稅": trade.get("securities_tax", 0),                    
                    "當日損益": trade.get("current_profit_loss", 0),
                    "持有天數": trade.get("holding_days", 0),
                    "出場原因": trade.get("exit_reason", ""), 
                    "開盤價": trade.get("open_price", 0),
                    "最高價": trade.get("high_price", 0),
                    "最低價": trade.get("low_price", 0),
                    "收盤價": trade.get("close_price", 0)
                })
                
            # 準備持有部位資料
            holding_data = []
            for position in complete_holding_positions:
                holding_data.append({
                    "進場日期": position.get("entry_date", ""),
                    "當前日期": position.get("current_date", ""),
                    "股票代碼": position.get("stock_id", ""),
                    "股票名稱": position.get("stock_name", ""),
                    "交易方向": "買入" if position.get("trade_direction", 1) == 1 else "賣出",
                    "進場價格": position.get("entry_price", 0),
                    "當前價格": position.get("current_price", 0),
                    "股數": position.get("shares", 0),
                    "未實現損益": position.get("unrealized_profit_loss", 0),
                    "未實現損益率(%)": position.get("unrealized_profit_loss_rate", 0),
                    "持有天數": position.get("holding_days", 0),
                    "出場價類型": position.get("exit_price_type", ""),
                    "當日進場價格": position.get("current_entry_price", 0),
                    "當日出場價格": position.get("current_exit_price", 0),
                    "當日損益": position.get("current_profit_loss", 0),
                    "當日損益率(%)": position.get("current_profit_loss_rate", 0),
                    "停利價格": position.get("take_profit_price", 0),
                    "停損價格": position.get("stop_loss_price", 0),
                    "開盤價": position.get("open_price", 0),
                    "最高價": position.get("high_price", 0),
                    "最低價": position.get("low_price", 0),
                    "收盤價": position.get("close_price", 0)
                })
            
            # 建立Excel檔案
            wb = Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 建立交易記錄工作表
            if trade_data:
                ws_trades = wb.create_sheet("交易記錄")
                
                # 設定標題樣式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 寫入標題
                headers = list(trade_data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws_trades.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 寫入資料
                for row, trade in enumerate(trade_data, 2):
                    for col, header in enumerate(headers, 1):
                        value = trade.get(header, "")
                        # 確保值不是空的字典或物件
                        if isinstance(value, dict) and not value:
                            value = ""
                        elif value is None:
                            value = ""
                        cell = ws_trades.cell(row=row, column=col, value=value)
                        
                        # 為損益欄位設定顏色
                        if header in ["報酬", "報酬率(%)", "淨損益", "未實現損益", "未實現損益率(%)", "當日損益", "當日損益率(%)"]:
                            if isinstance(value, (int, float)) and value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif isinstance(value, (int, float)) and value < 0:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
            # 建立持有部位工作表
            if holding_data:
                ws_holdings = wb.create_sheet("持有部位")
                
                # 設定標題樣式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 寫入標題
                headers = list(holding_data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws_holdings.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 寫入資料
                for row, position in enumerate(holding_data, 2):
                    for col, header in enumerate(headers, 1):
                        value = position.get(header, "")
                        # 確保值不是空的字典或物件
                        if isinstance(value, dict) and not value:
                            value = ""
                        elif value is None:
                            value = ""
                        cell = ws_holdings.cell(row=row, column=col, value=value)
                        
                        # 為損益欄位設定顏色
                        if header in ["未實現損益", "未實現損益率(%)", "當日損益", "當日損益率(%)"]:
                            if isinstance(value, (int, float)) and value > 0:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            elif isinstance(value, (int, float)) and value < 0:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # 調整欄寬
                for col in ws_holdings.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_holdings.column_dimensions[column].width = adjusted_width
            
            # 建立摘要工作表
            ws_summary = wb.create_sheet("回測摘要")
            
            # 設定標題樣式
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True)
            
            # 寫入摘要資訊
            ws_summary.cell(row=1, column=1, value="回測摘要").font = title_font
            ws_summary.cell(row=3, column=1, value="策略名稱").font = header_font
            ws_summary.cell(row=3, column=2, value=strategy_name)
            ws_summary.cell(row=4, column=1, value="匯出時間").font = header_font
            ws_summary.cell(row=4, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ws_summary.cell(row=5, column=1, value="交易記錄數量").font = header_font
            ws_summary.cell(row=5, column=2, value=len(trade_data))
            ws_summary.cell(row=6, column=1, value="持有部位數量").font = header_font
            ws_summary.cell(row=6, column=2, value=len(holding_data))
            
            # 調整欄寬
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 30
            
            # 將工作表順序調整為：摘要、交易記錄、持有部位
            wb._sheets = [ws_summary] + [ws for ws in wb._sheets if ws != ws_summary]
            
            # 儲存到記憶體
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 產生檔案名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"detailed_backtest_{strategy_name}_{timestamp}.xlsx"
            
            return StreamingResponse(
                BytesIO(output.getvalue()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except Exception as e:
            print_log(f"export_detailed_records error: {e}")
            raise HTTPException(status_code=500, detail=str(e))
    
    @staticmethod
    async def get_excel_format() -> Dict[str, Any]:
        """
        取得Excel檔案格式說明
        
        Returns:
            格式說明字典
        """
        try:
            format_info = {
                "success": True,
                "format": {
                    "api_mode": {
                        "description": "API模式：Excel檔案需包含股票代碼和日期",
                        "columns": ["stock_id", "date"],
                        "example": [
                            {"stock_id": "2330", "date": "2024-01-01"},
                            {"stock_id": "2330", "date": "2024-01-02"}
                        ]
                    },
                    "excel_mode": {
                        "description": "Excel模式：Excel檔案需包含完整的K線資料",
                        "columns": ["stock_id", "date", "open", "high", "low", "close"],
                        "example": [
                            {"stock_id": "2330", "date": "2024-01-01", "open": 100, "high": 105, "low": 98, "close": 103},
                            {"stock_id": "2330", "date": "2024-01-02", "open": 103, "high": 108, "low": 102, "close": 106}
                        ]
                    },
                    "bookbuilding": {
                        "description": "詢圈公告策略：Excel檔案只需包含股票代碼",
                        "columns": ["stock_id"],
                        "example": [
                            {"stock_id": "2330"},
                            {"stock_id": "2317"}
                        ]
                    }
                }
            }
            return format_info
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
    
    @staticmethod
    async def get_complete_trade_data(trade_records: List[Dict], holding_positions: List[Dict]) -> Dict[str, Any]:
        """
        取得完整的交易記錄資料，包含所有欄位
        
        Args:
            trade_records: 交易記錄列表
            holding_positions: 持有部位列表
            
        Returns:
            完整的交易記錄和持有部位資料
        """
        try:
            # 輔助函數：安全地取得值
            def safe_get_value(obj, key, default=""):
                """安全地從物件或字典中取得值"""
                try:
                    if hasattr(obj, '__dict__'):
                        # 如果是物件
                        value = getattr(obj, key, default)
                    else:
                        # 如果是字典
                        value = obj.get(key, default)
                    
                    # 處理特殊值
                    if value is None:
                        return default
                    elif isinstance(value, dict) and not value:
                        return default
                    elif isinstance(value, (list, tuple)) and not value:
                        return default
                    else:
                        return value
                except:
                    return default
            
            # 轉換交易記錄為完整字典格式
            complete_trade_records = []
            for trade in trade_records:
                complete_trade = {
                    "position_id": safe_get_value(trade, "position_id", ""),
                    "entry_date": safe_get_value(trade, "entry_date", ""),
                    "exit_date": safe_get_value(trade, "exit_date", ""),
                    "stock_id": safe_get_value(trade, "stock_id", ""),
                    "stock_name": safe_get_value(trade, "stock_name", ""),
                    "trade_direction": safe_get_value(trade, "trade_direction", 1),
                    "entry_price": safe_get_value(trade, "entry_price", 0),
                    "exit_price": safe_get_value(trade, "exit_price", 0),
                    "shares": safe_get_value(trade, "shares", 0),
                    "profit_loss": safe_get_value(trade, "profit_loss", safe_get_value(trade, "profit", 0)),
                    "profit_loss_rate": safe_get_value(trade, "profit_loss_rate", safe_get_value(trade, "profit_rate", 0)),
                    "commission": safe_get_value(trade, "commission", 0),
                    "securities_tax": safe_get_value(trade, "securities_tax", 0),
                    "net_profit_loss": safe_get_value(trade, "net_profit_loss", safe_get_value(trade, "profit", 0)),
                    "holding_days": safe_get_value(trade, "holding_days", 0),
                    "exit_reason": safe_get_value(trade, "exit_reason", safe_get_value(trade, "exit_status", "")),
                    "current_price": safe_get_value(trade, "current_price", 0),
                    "unrealized_profit_loss": safe_get_value(trade, "unrealized_profit_loss", 0),
                    "unrealized_profit_loss_rate": safe_get_value(trade, "unrealized_profit_loss_rate", 0),
                    "current_date": safe_get_value(trade, "current_date", ""),
                    "exit_price_type": safe_get_value(trade, "exit_price_type", ""),
                    "current_entry_price": safe_get_value(trade, "current_entry_price", 0),
                    "current_exit_price": safe_get_value(trade, "current_exit_price", 0),
                    "current_profit_loss": safe_get_value(trade, "current_profit_loss", 0),
                    "current_profit_loss_rate": safe_get_value(trade, "current_profit_loss_rate", 0),
                    "take_profit_price": safe_get_value(trade, "take_profit_price", 0),
                    "stop_loss_price": safe_get_value(trade, "stop_loss_price", 0),
                    "open_price": safe_get_value(trade, "open_price", 0),
                    "high_price": safe_get_value(trade, "high_price", 0),
                    "low_price": safe_get_value(trade, "low_price", 0),
                    "close_price": safe_get_value(trade, "close_price", 0)
                }
                complete_trade_records.append(complete_trade)
            
            # 轉換持有部位為完整字典格式
            complete_holding_positions = []
            for position in holding_positions:
                complete_position = {
                    "position_id": safe_get_value(position, "position_id", ""),
                    "entry_date": safe_get_value(position, "entry_date", ""),
                    "current_date": safe_get_value(position, "current_date", ""),
                    "stock_id": safe_get_value(position, "stock_id", ""),
                    "stock_name": safe_get_value(position, "stock_name", ""),
                    "trade_direction": safe_get_value(position, "trade_direction", 1),
                    "entry_price": safe_get_value(position, "entry_price", 0),
                    "current_price": safe_get_value(position, "current_price", 0),
                    "shares": safe_get_value(position, "shares", 0),
                    "unrealized_profit_loss": safe_get_value(position, "unrealized_profit_loss", 0),
                    "unrealized_profit_loss_rate": safe_get_value(position, "unrealized_profit_loss_rate", 0),
                    "holding_days": safe_get_value(position, "holding_days", 0),
                    "exit_price_type": safe_get_value(position, "exit_price_type", ""),
                    "current_entry_price": safe_get_value(position, "current_entry_price", 0),
                    "current_exit_price": safe_get_value(position, "current_exit_price", 0),
                    "current_profit_loss": safe_get_value(position, "current_profit_loss", 0),
                    "current_profit_loss_rate": safe_get_value(position, "current_profit_loss_rate", 0),
                    "take_profit_price": safe_get_value(position, "take_profit_price", 0),
                    "stop_loss_price": safe_get_value(position, "stop_loss_price", 0),
                    "open_price": safe_get_value(position, "open_price", 0),
                    "high_price": safe_get_value(position, "high_price", 0),
                    "low_price": safe_get_value(position, "low_price", 0),
                    "close_price": safe_get_value(position, "close_price", 0)
                }
                complete_holding_positions.append(complete_position)
            
            print_log(f" -------- 完整持有部位欄位數量: {len(complete_holding_positions[0].keys()) if complete_holding_positions else 0}")
            print_log(f" -------- 完整持有部位欄位: {list(complete_holding_positions[0].keys()) if complete_holding_positions else []}")
            
            return {
                "success": True,
                "complete_trade_records": complete_trade_records,
                "complete_holding_positions": complete_holding_positions
            }
            
        except Exception as e:
            print_log(f"get_complete_trade_data error: {e}")
            raise HTTPException(status_code=500, detail=str(e)) 